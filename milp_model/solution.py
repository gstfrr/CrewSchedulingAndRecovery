# -*- coding: utf-8 -*-
"""Solution File

This file is used to create and define useful functions to help the analysis of the solution. Here we can find
functions to calculate statistics and write files.
"""

from gurobipy import Model, GRB, Var, LinExpr
import pandas as pd
from openpyxl import load_workbook


def clean_model(m: Model) -> None:
    """From https://support.gurobi.com/hc/en-us/community/posts/360048150752-Remove-variables-with-0-coefficient-in-the-LP-file
    This function removes variables that are not used in the model. It is useful to reduce the size of the model.

    :param m: Model: Gurobi model to have the unused variables removed.

    """
    m.update()

    gcfuncs = {
        GRB.GENCONSTR_MAX: m.getGenConstrMax,
        GRB.GENCONSTR_MIN: m.getGenConstrMin,
        GRB.GENCONSTR_ABS: m.getGenConstrAbs,
        GRB.GENCONSTR_AND: m.getGenConstrAnd,
        GRB.GENCONSTR_OR: m.getGenConstrOr,
        GRB.GENCONSTR_NORM: m.getGenConstrNorm,
        GRB.GENCONSTR_INDICATOR: m.getGenConstrIndicator,
        GRB.GENCONSTR_PWL: m.getGenConstrPWL,
        GRB.GENCONSTR_POLY: m.getGenConstrPoly,
        GRB.GENCONSTR_EXP: m.getGenConstrExp,
        GRB.GENCONSTR_EXPA: m.getGenConstrExpA,
        GRB.GENCONSTR_LOG: m.getGenConstrLog,
        GRB.GENCONSTR_LOGA: m.getGenConstrLogA,
        GRB.GENCONSTR_LOGISTIC: m.getGenConstrLogistic,
        GRB.GENCONSTR_POW: m.getGenConstrPow,
        GRB.GENCONSTR_SIN: m.getGenConstrSin,
        GRB.GENCONSTR_COS: m.getGenConstrCos,
        GRB.GENCONSTR_TAN: m.getGenConstrTan,
    }

    # Indices of variables participating in general constraints
    gcvars = set()

    for gc in m.getGenConstrs():
        retvals = gcfuncs[gc.GenConstrType](gc)

        # Vars are found in return values of type Var, LinExpr, and List[Var]
        for rv in retvals:
            if isinstance(rv, Var):
                gcvars.add(rv.index)
            elif isinstance(rv, LinExpr):
                for i in range(rv.size()):
                    gcvars.add(rv.getVar(i).index)
            elif isinstance(rv, list) and len(rv) > 0 and isinstance(rv[0], Var):
                for v in rv:
                    gcvars.add(v.index)

    to_remove = [v for v in m.getVars() if not m.getCol(v).size() and not v.Obj and v.index not in gcvars]
    m.remove(to_remove)
    print(f"Removed {len(to_remove)} unused variables")


def get_var_details(var: Var) -> list[object]:
    """Returns the specified attributes of a variable object.

    :param var: Var: Optimized variable.

    """
    return [var.X, var.Obj, var.VarName, var.LB, var.UB]


def get_var_columns() -> list[str]:
    """ """
    return ['Value', 'ObjCoef', 'VarName', 'LB', 'UB']


def write_solution(flight_pilot_assignment_vars, filename):
    rows = []
    for value in flight_pilot_assignment_vars.values():
        var = value.variable
        if var.X == 0:
            continue
        rows.append(
            [value.flight.name, value.pilot.name, value.flight.duration, value.flight.start, value.flight.end] +
            get_var_details(var))

    df = pd.DataFrame(rows, columns=['Flight', 'Pilot', 'Duration', 'Start', 'End'] + get_var_columns())

    with pd.ExcelWriter(filename) as writer:
        df.to_excel(writer, sheet_name="Flights", index=False)

    df.to_csv('output/solution.csv', index=False)

    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)

    return df
